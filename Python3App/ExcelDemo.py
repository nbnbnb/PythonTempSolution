from openpyxl import load_workbook

def main():    
    hf()  # 代驾
    sd()  # 自驾

def hf():
    res = []
    wb = load_workbook(filename=r'x:\HF.xlsx',read_only=True)
    ws = wb.active
    for row in ws.iter_rows(row_offset=1):      
        orderType = str(row[1].value).strip()
        orderStatus = str(row[2].value).strip()
        driverType = str(row[5].value).strip()
        reAssign = str(row[6].value).strip()
        question = str(row[7].value).strip().replace('\n',' ')
        answer = str(row[10].value).strip().replace('\n',' ')

        for payStatus in str(row[6].value).strip().split('/'):
            for time in str(row[4].value).strip().split('/'):
                status = getOrderType(orderType) + getOrderStatus(orderStatus) + getPayStatus(payStatus) + getBeforeMiddleAfter(time) + getDriverType(driverType) + getReAssign(reAssign)
                desc = orderType + '-' + orderStatus + '-' + payStatus + '-' + time + '-' + driverType + '-' + reAssign
                if answer != 'None':
                    res.append(status + '\t' + desc + '\t' + question + '\t' + answer)

    with open(r'x:\HF.txt','+w') as f:
             for line in res:
                 f.write(line + '\n')

def sd():
    res = []
    wb = load_workbook(filename=r'x:\SD.xlsx',read_only=True)
    ws = wb.active
    for row in ws.iter_rows(row_offset=1):      
        orderType = str(row[1].value).strip()
        orderStatus = str(row[3].value).strip()
        question = str(row[6].value).strip().replace('\n',' ')
        answer = str(row[8].value).strip().replace('\n',' ')

        for payType in str(row[2].value).strip().split('/'):
            for payStatus in str(row[4].value).strip().split('/'):
                for time in str(row[5].value).strip().split('/'):
                    status = getOrderType(orderType) + getPayType(payType) + getOrderStatus(orderStatus) + getPayStatus(payStatus) + getBeforeMiddleAfter(time) 
                    desc = orderType + '-' + payType + '-' + orderStatus + '-' + payStatus + '-' + time 
                    if answer != 'None':
                        res.append(status + '\t' + desc + '\t' + question + '\t' + answer)

    with open(r'x:\SD.txt','+w') as f:
             for line in res:
                 f.write(line + '\n')

def getBase(val,mapping):
    try:
        index = mapping.index(val)
        return str(index)
    except ValueError:
        return '0'    

def getOrderType(val):
    mapping = ['','国内接送机','国内日租','海外接送机','海外日租','接送火车','随叫随到','国内自驾','海外自驾']
    return getBase(val,mapping)

def getOrderStatus(val):
    mapping = ['','已确认产品和库存','已下单','已确认客户及扣款','取消','完成']
    return getBase(val,mapping)

def getPayStatus(val):
    mapping = ['','已确认产品和库存','已下单','已确认客户及扣款','取消','完成']
    return getBase(val,mapping)

def getBeforeMiddleAfter(val):
    mapping = ['','前','中','后']
    return getBase(val,mapping)

def getDriverType(val):
    mapping = ['','调度电话','司机信息']
    return getBase(val,mapping)

def getReAssign(val):
    mapping = ['','否','是']
    return getBase(val,mapping)

def getPayType(val):
    mapping = ['','现付','预付','预付定金','供应商收款']
    return getBase(val,mapping)

